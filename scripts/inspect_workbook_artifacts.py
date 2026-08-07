from __future__ import annotations

import json
import sys
import zipfile
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


for argument in sys.argv[1:]:
    path = Path(argument)
    workbook = load_workbook(path, read_only=True, data_only=False, keep_links=True)
    sheets = []
    for sheet in workbook.worksheets:
        style_counts = Counter()
        formula_columns = Counter()
        comment_count = 0
        hyperlink_count = 0
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    style_counts[getattr(cell, "style_id", getattr(cell, "_style_id", None))] += 1
                if cell.data_type == "f":
                    formula_columns[cell.column_letter] += 1
                if getattr(cell, "comment", None):
                    comment_count += 1
                if getattr(cell, "hyperlink", None):
                    hyperlink_count += 1
        sheets.append({
            "name": sheet.title,
            "state": sheet.sheet_state,
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "auto_filter": getattr(getattr(sheet, "auto_filter", None), "ref", None),
            "freeze_panes": str(getattr(sheet, "freeze_panes", "") or ""),
            "merged_ranges": [str(item) for item in getattr(getattr(sheet, "merged_cells", None), "ranges", [])],
            "formula_columns": formula_columns,
            "style_counts": style_counts,
            "comments": comment_count,
            "hyperlinks": hyperlink_count,
        })
    workbook.close()
    with zipfile.ZipFile(path) as archive:
        names = archive.namelist()
        package_artifacts = {
            "external_links": [name for name in names if name.startswith("xl/externalLinks/")],
            "comments": [name for name in names if "/comments" in name],
            "connections": [name for name in names if name == "xl/connections.xml"],
            "pivot_tables": [name for name in names if name.startswith("xl/pivotTables/")],
            "tables": [name for name in names if name.startswith("xl/tables/")],
        }
    print(json.dumps({"file": str(path), "sheets": sheets, "package": package_artifacts}, ensure_ascii=False, indent=2, default=dict))
