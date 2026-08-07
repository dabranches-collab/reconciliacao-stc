from __future__ import annotations

import json
import sys
from collections import Counter, defaultdict
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook


def cents(value) -> int:
    return int((Decimal(str(value or 0)) * 100).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def date_key(value) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value or "").strip()
    for fmt in ("%d-%m-%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return text


def load_stc(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["STC"]
    records = []
    for row_number, row in enumerate(ws.iter_rows(min_row=17, values_only=True), start=17):
        if not any(value is not None for value in row):
            continue
        records.append({
            "row": row_number,
            "date": date_key(row[2]),
            "dc": row[3],
            "amount": cents(row[4]),
            "operation": str(row[6]).strip() if row[6] is not None else "",
            "description": str(row[7] or "").strip(),
            "obs": str(row[9] or "").strip(),
            "do": str(row[10] or "").strip(),
            "ordering_party": str(row[11] or "").strip(),
            "beneficiary": str(row[12] or "").strip(),
            "iban": str(row[13] or "").strip(),
            "bic": str(row[14] or "").strip(),
        })
    wb.close()
    return records


def load_extract(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Folha1"]
    records = []
    for row_number, row in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
        if not any(value is not None for value in row):
            continue
        debit = cents(row[5])
        credit = cents(row[6])
        records.append({
            "row": row_number,
            "date": date_key(row[1]),
            "amount": debit - credit,
            "operation": str(row[4]).strip() if row[4] is not None else "",
            "description": str(row[3] or "").strip(),
            "debit": debit,
            "credit": credit,
            "origin": str(row[7] or "").strip(),
            "branch": str(row[8] or "").strip(),
            "currency": str(row[9] or "").strip(),
        })
    wb.close()
    return records


def load_appended_rec(path: Path, start_row: int):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Rec."]
    records = []
    for row_number, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
        if not any(value is not None for value in row):
            continue
        records.append({
            "row": row_number,
            "date": date_key(row[2]),
            "dc": row[3],
            "amount": cents(row[4]),
            "operation": str(row[6]).strip() if row[6] is not None else "",
            "description": str(row[7] or "").strip(),
            "obs": str(row[8] or "").strip(),
            "do": str(row[9] or "").strip(),
        })
    wb.close()
    return records


def key(record):
    return record["date"], record["operation"], record["amount"]


def consume(source, target):
    available = defaultdict(list)
    for record in source:
        available[key(record)].append(record)
    matched = []
    unmatched = []
    for record in target:
        bucket = available.get(key(record))
        if bucket:
            matched.append((bucket.pop(), record))
        else:
            unmatched.append(record)
    remaining = [record for bucket in available.values() for record in bucket]
    return matched, unmatched, remaining


previous_path, extract_path, final_path = map(Path, sys.argv[1:4])
previous = load_stc(previous_path)
new = load_extract(extract_path)
# The Banka report ends with a totals row that has amounts but no movement identity.
new = [record for record in new if record["date"] and record["operation"]]
final = load_stc(final_path)
appended = load_appended_rec(final_path, 375280)

universe = [{**record, "source": "previous"} for record in previous] + [
    {**record, "source": "new"} for record in new
]
pending_matches, pending_unmatched, closed = consume(universe, final)
closed_matches, rec_unmatched, not_in_rec = consume(closed, appended)

groups = defaultdict(list)
for record in appended:
    groups[record["operation"]].append(record)

group_summaries = []
for operation, records in groups.items():
    values = [record["amount"] for record in records]
    dates = sorted({record["date"] for record in records})
    group_summaries.append({
        "operation": operation,
        "count": len(records),
        "sum_cents": sum(values),
        "dates": dates,
        "positive": sum(value > 0 for value in values),
        "negative": sum(value < 0 for value in values),
        "descriptions": sorted({record["description"] for record in records}),
    })

source_by_key = defaultdict(list)
for record in closed:
    source_by_key[key(record)].append(record["source"])
group_sources = Counter()
for operation, records in groups.items():
    sources = []
    local = {k: list(v) for k, v in source_by_key.items()}
    for record in records:
        bucket = local.get(key(record), [])
        sources.append(bucket.pop() if bucket else "unknown")
    group_sources["+".join(sorted(set(sources)))] += 1

summary = {
    "counts": {
        "previous_pending": len(previous),
        "new_extract": len(new),
        "final_pending": len(final),
        "appended_reconciled": len(appended),
        "universe": len(universe),
        "identity_difference": len(universe) - len(final) - len(appended),
        "final_pending_matched": len(pending_matches),
        "final_pending_unmatched": len(pending_unmatched),
        "closed_matched_to_appended": len(closed_matches),
        "appended_unmatched": len(rec_unmatched),
        "closed_not_in_appended": len(not_in_rec),
    },
    "pending_sources": Counter(source["source"] for source, _ in pending_matches),
    "reconciled_sources": Counter(source["source"] for source, _ in closed_matches),
    "groups": {
        "total": len(group_summaries),
        "zero_sum": sum(group["sum_cents"] == 0 for group in group_summaries),
        "nonzero_sum": sum(group["sum_cents"] != 0 for group in group_summaries),
        "size_distribution": Counter(group["count"] for group in group_summaries),
        "source_composition": group_sources,
        "multi_date": sum(len(group["dates"]) > 1 for group in group_summaries),
        "single_sided": sum(group["positive"] == 0 or group["negative"] == 0 for group in group_summaries),
    },
    "descriptions": {
        "previous_pending": Counter(record["description"] for record in previous).most_common(),
        "new_extract": Counter(record["description"] for record in new).most_common(),
        "final_pending": Counter(record["description"] for record in final).most_common(),
        "appended_reconciled": Counter(record["description"] for record in appended).most_common(),
    },
    "by_date": {},
    "by_sign": {},
    "nonzero_group_samples": [group for group in group_summaries if group["sum_cents"] != 0][:20],
    "unmatched_samples": {
        "pending": pending_unmatched[:20],
        "appended": rec_unmatched[:20],
        "closed": not_in_rec[:20],
    },
}

for label, records in (("previous", previous), ("new", new), ("final", final), ("appended", appended)):
    date_stats = defaultdict(lambda: {"count": 0, "sum_cents": 0, "positive": 0, "negative": 0})
    for record in records:
        item = date_stats[record["date"]]
        item["count"] += 1
        item["sum_cents"] += record["amount"]
        item["positive"] += record["amount"] > 0
        item["negative"] += record["amount"] < 0
    summary["by_date"][label] = dict(sorted(date_stats.items()))
    summary["by_sign"][label] = {
        "positive_count": sum(record["amount"] > 0 for record in records),
        "positive_sum_cents": sum(record["amount"] for record in records if record["amount"] > 0),
        "negative_count": sum(record["amount"] < 0 for record in records),
        "negative_sum_cents": sum(record["amount"] for record in records if record["amount"] < 0),
        "total_cents": sum(record["amount"] for record in records),
    }

positive_by_abs = defaultdict(list)
negative_by_abs = defaultdict(list)
for record in appended:
    if record["amount"] > 0:
        positive_by_abs[record["amount"]].append(record)
    elif record["amount"] < 0:
        negative_by_abs[-record["amount"]].append(record)
paired_lines = sum(2 * min(len(positive_by_abs[value]), len(negative_by_abs[value])) for value in positive_by_abs)
unambiguous_lines = sum(
    2 for value in positive_by_abs
    if len(positive_by_abs[value]) == 1 and len(negative_by_abs[value]) == 1
)
summary["exact_opposite_values"] = {
    "paired_lines_maximum": paired_lines,
    "unambiguous_lines": unambiguous_lines,
    "coverage_percent": round(100 * paired_lines / len(appended), 4) if appended else 0,
}
print(json.dumps(summary, ensure_ascii=False, indent=2, default=dict))
