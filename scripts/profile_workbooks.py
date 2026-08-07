from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def scalar(value):
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def profile(path: Path):
    workbook = load_workbook(path, read_only=True, data_only=False, keep_links=True)
    result = {"file": str(path), "sheets": []}
    for sheet in workbook.worksheets:
        sample = []
        formulas = []
        nonempty = 0
        for row_index, row in enumerate(sheet.iter_rows(), start=1):
            values = [scalar(cell.value) for cell in row]
            nonempty += sum(value is not None for value in values)
            if row_index <= 15:
                sample.append(values)
            for cell in row:
                if cell.data_type == "f":
                    if len(formulas) < 100:
                        formulas.append({"cell": cell.coordinate, "formula": cell.value})
        result["sheets"].append({
            "name": sheet.title,
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "nonempty_cells": nonempty,
            "sample": sample,
            "formula_count_or_more": len(formulas),
            "formula_samples": formulas,
        })
    workbook.close()
    return result


for argument in sys.argv[1:]:
    print(json.dumps(profile(Path(argument)), ensure_ascii=False))
