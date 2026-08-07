from __future__ import annotations

import hashlib
import json
import sys
from collections import Counter
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook


def cents(value) -> int:
    return int((Decimal(str(value or 0)) * 100).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def date_key(value) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    return datetime.strptime(str(value).strip(), "%d-%m-%Y").date().isoformat()


source = Path(sys.argv[1])
destination = Path(sys.argv[2])
file_hash = hashlib.sha256(source.read_bytes()).hexdigest()
workbook = load_workbook(source, read_only=True, data_only=True)
sheet = workbook["STC"]
occurrences: Counter[str] = Counter()
movements = []

for row_number, row in enumerate(sheet.iter_rows(min_row=17, values_only=True), start=17):
    if not any(value is not None for value in row):
        continue
    movement = {
        "source_row": row_number,
        "movement_date": date_key(row[2]),
        "dc": str(row[3] or "").strip(),
        "amount_minor": cents(row[4]),
        "operation": str(row[6] or "").strip(),
        "description": str(row[7] or "").strip(),
        "observation": str(row[9] or "").strip(),
        "document_number": str(row[10] or "").strip(),
        "ordering_party": str(row[11] or "").strip(),
        "beneficiary": str(row[12] or "").strip(),
        "iban": str(row[13] or "").strip(),
        "bic": str(row[14] or "").strip(),
        "state": "open",
    }
    canonical = json.dumps({k: v for k, v in movement.items() if k != "source_row"}, sort_keys=True, ensure_ascii=False)
    occurrences[canonical] += 1
    movement["fingerprint"] = hashlib.sha256(f"{canonical}|{occurrences[canonical]}".encode("utf-8")).hexdigest()
    movements.append(movement)

workbook.close()
assert len(movements) == 686, len(movements)
assert sum(item["amount_minor"] for item in movements) == -482_824_159_179

payload = {
    "source": {"name": source.name, "sha256": file_hash},
    "position": {
        "position_date": "2026-08-06",
        "previous_pending_count": 1320,
        "new_movement_count": 11355,
        "reconciled_count": 11989,
        "closing_pending_count": 686,
        "accounting_balance_minor": -482_824_159_179,
        "closing_pending_balance_minor": -482_824_159_179,
        "status": "validated",
    },
    "groups": [5938, 3050, 2802, 195, 2, 2],
    "movements": movements,
}
destination.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
print(json.dumps({"count": len(movements), "sum_minor": sum(item["amount_minor"] for item in movements), "output": str(destination)}))
