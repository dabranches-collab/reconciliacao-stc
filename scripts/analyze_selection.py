from __future__ import annotations

import json
import sys
from collections import Counter, defaultdict, deque
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook


def cents(value):
    return int((Decimal(str(value or 0)) * 100).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def date_key(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value or "").strip()
    for fmt in ("%d-%m-%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return text


def load_stc(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["STC"]
    records = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=17, values_only=True), start=17):
        if any(value is not None for value in row):
            records.append({
                "row": row_number,
                "date": date_key(row[2]),
                "amount": cents(row[4]),
                "operation": str(row[6]).strip() if row[6] is not None else "",
                "description": str(row[7] or "").strip(),
            })
    workbook.close()
    return records


def load_extract(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Folha1"]
    records = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=5, values_only=True), start=5):
        if any(value is not None for value in row):
            records.append({
                "row": row_number,
                "date": date_key(row[1]),
                "amount": cents(row[5]) - cents(row[6]),
                "operation": str(row[4]).strip() if row[4] is not None else "",
                "description": str(row[3] or "").strip(),
            })
    workbook.close()
    return records


def key(record):
    return record["date"], record["operation"], record["amount"]


previous_path, extract_path, final_path = map(Path, sys.argv[1:4])
previous = load_stc(previous_path)
new = [row for row in load_extract(extract_path) if row["date"] and row["operation"]]
final = load_stc(final_path)

pending_keys = defaultdict(deque)
for row in final:
    pending_keys[key(row)].append(row)

ordered = []
for row in new:
    is_pending = bool(pending_keys[key(row)])
    if is_pending:
        pending_keys[key(row)].popleft()
    ordered.append({**row, "status": "pending" if is_pending else "closed"})

runs = []
for row in ordered:
    if not runs or runs[-1]["status"] != row["status"]:
        runs.append({
            "status": row["status"],
            "start_row": row["row"],
            "end_row": row["row"],
            "count": 1,
            "sum_cents": row["amount"],
            "start_date": row["date"],
            "end_date": row["date"],
        })
    else:
        run = runs[-1]
        run["end_row"] = row["row"]
        run["count"] += 1
        run["sum_cents"] += row["amount"]
        run["end_date"] = row["date"]

combined_running = sum(row["amount"] for row in previous)
zero_boundaries = []
closest = None
for index, row in enumerate(ordered):
    combined_running += row["amount"]
    absolute = abs(combined_running)
    if closest is None or absolute < closest["absolute_cents"]:
        closest = {
            "extract_index": index,
            "row": row["row"],
            "date": row["date"],
            "operation": row["operation"],
            "balance_cents": combined_running,
            "absolute_cents": absolute,
        }
    if combined_running == 0:
        zero_boundaries.append({
            "extract_index": index,
            "row": row["row"],
            "date": row["date"],
            "operation": row["operation"],
        })

pending_positions = [index for index, row in enumerate(ordered) if row["status"] == "pending"]
closed_positions = [index for index, row in enumerate(ordered) if row["status"] == "closed"]
first_pending = min(pending_positions)
last_closed = max(closed_positions)

def description_stats(status):
    records = [row for row in ordered if row["status"] == status and row["date"] == "2026-08-06"]
    return Counter(row["description"] for row in records)

summary = {
    "runs": runs,
    "first_pending": ordered[first_pending],
    "last_closed": ordered[last_closed],
    "pending_is_exact_suffix": first_pending == last_closed + 1,
    "zero_boundaries": zero_boundaries,
    "closest_running_balance": closest,
    "august_6_descriptions": {
        "closed": description_stats("closed"),
        "pending": description_stats("pending"),
    },
    "pending_order_matches_extract": [key(row) for row in final] == [key(row) for row in ordered if row["status"] == "pending"],
}

closed_balance = sum(row["amount"] for row in previous)
closed_boundaries = []
segment_start = 0
closed_rows = [row for row in ordered if row["status"] == "closed"]
for index, row in enumerate(closed_rows):
    closed_balance += row["amount"]
    if closed_balance == 0:
        segment = closed_rows[segment_start:index + 1]
        closed_boundaries.append({
            "end_extract_row": row["row"],
            "date": row["date"],
            "new_movement_count": len(segment),
            "previous_movement_count": len(previous) if segment_start == 0 else 0,
            "descriptions": Counter(item["description"] for item in segment),
            "positive_operations": Counter(item["operation"] for item in segment if item["amount"] > 0).most_common(10),
        })
        segment_start = index + 1
summary["closed_sequence_zero_boundaries"] = closed_boundaries

description_runs = []
for row in ordered:
    label = (row["date"], row["description"], row["status"], "positive" if row["amount"] > 0 else "negative")
    if not description_runs or description_runs[-1]["label"] != label:
        description_runs.append({
            "label": label,
            "start_row": row["row"],
            "end_row": row["row"],
            "count": 1,
            "sum_cents": row["amount"],
        })
    else:
        description_runs[-1]["end_row"] = row["row"]
        description_runs[-1]["count"] += 1
        description_runs[-1]["sum_cents"] += row["amount"]
summary["august_6_description_runs"] = [run for run in description_runs if run["label"][0] == "2026-08-06"]
print(json.dumps(summary, ensure_ascii=False, indent=2, default=dict))
