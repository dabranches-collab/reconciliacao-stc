from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


path = Path(sys.argv[1])
sheet_name = sys.argv[2]
start = int(sys.argv[3])
end = int(sys.argv[4])
workbook = load_workbook(path, read_only=True, data_only=True)
sheet = workbook[sheet_name]
for number, row in enumerate(
    sheet.iter_rows(min_row=start, max_row=end, values_only=True), start=start
):
    values = [value.isoformat() if hasattr(value, "isoformat") else value for value in row]
    print(json.dumps({"row": number, "values": values}, ensure_ascii=False))
workbook.close()
